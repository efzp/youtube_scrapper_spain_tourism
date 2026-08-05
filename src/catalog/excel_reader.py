"""Lectura de la tabla estructurada `tbl_lugares` de Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


class CatalogReadError(ValueError):
    """Indica que el libro no contiene la tabla o estructura esperada."""


def read_excel_table(path: Path, table_name: str = "tbl_lugares") -> list[dict[str, Any]]:
    """Lee una tabla por nombre, sin depender del nombre de la hoja."""
    if not path.is_file():
        raise CatalogReadError(f"No existe el catalogo configurado: {path}")
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            if table_name not in worksheet.tables:
                continue
            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            headers = [
                worksheet.cell(row=min_row, column=column).value
                for column in range(min_col, max_col + 1)
            ]
            if any(not isinstance(header, str) or not header.strip() for header in headers):
                raise CatalogReadError(f"La tabla {table_name} tiene encabezados vacios.")
            return [
                dict(
                    zip(
                        headers,
                        (
                            worksheet.cell(row=row, column=column).value
                            for column in range(min_col, max_col + 1)
                        ),
                        strict=True,
                    )
                )
                for row in range(min_row + 1, max_row + 1)
            ]
    finally:
        workbook.close()
    raise CatalogReadError(f"No se encontro la tabla {table_name!r} en {path}.")

